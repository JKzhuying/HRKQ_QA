import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os


def is_yellow_fill(cell):
    """判断单元格是否为黄色填充"""
    fill = cell.fill
    if fill and fill.fill_type == 'solid' and fill.fgColor:
        # 获取颜色值 (可能是 AARRGGBB 或 RRGGBB 格式)
        color = fill.fgColor.rgb
        if color and isinstance(color, str) and len(color) >= 6:
            try:
                # 处理 AARRGGBB 格式 (跳过alpha通道)
                if len(color) == 8:
                    r = int(color[2:4], 16)
                    g = int(color[4:6], 16)
                    b = int(color[6:8], 16)
                else:
                    r = int(color[0:2], 16)
                    g = int(color[2:4], 16)
                    b = int(color[4:6], 16)
                # 黄色: R>200, G>200, B<100
                if r > 200 and g > 200 and b < 100:
                    return True
            except ValueError:
                pass
            # 检查常见的黄色颜色值（包含alpha通道）
            yellow_colors = ['00FFFF00', 'FFFFFF00', 'FFFF00B0', 
                           'FFFFFFC0', 'FFFFEB9C', 'FFFFFF99',
                           'FFFFFF69', 'FFFFFF0F', 'FFFFFF88']
            color_upper = color.upper()
            # 检查后6位（RGB部分）
            rgb_part = color_upper[-6:]
            if rgb_part == 'FFFF00' or rgb_part in [c[-6:] for c in yellow_colors]:
                return True
            if color_upper in yellow_colors:
                return True
        # 检查主题色中的黄色（主题色5或6通常是黄色）
        if fill.fgColor.theme is not None:
            if fill.fgColor.theme in [5, 6]:
                return True
    return False


def parse_date(date_value):
    """解析日期"""
    if date_value is None:
        return None
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')
    if isinstance(date_value, str):
        # 尝试多种格式
        formats = ['%Y/%m/%d', '%Y-%m-%d', '%Y%m%d', '%m/%d/%Y']
        for fmt in formats:
            try:
                dt = datetime.strptime(date_value.strip(), fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def parse_amount(amount_value):
    """解析金额"""
    if amount_value is None:
        return 0.0
    if isinstance(amount_value, (int, float)):
        return float(amount_value)
    if isinstance(amount_value, str):
        try:
            return float(amount_value.strip().replace(',', ''))
        except ValueError:
            return 0.0
    return 0.0


def detect_payment_method(item_name):
    """根据项目名称检测支付方式"""
    if not item_name:
        return 'cash'
    item_name = str(item_name)
    if '支付宝' in item_name or 'AliPay' in item_name or 'alipay' in item_name.lower():
        return 'alipay'
    elif '微信' in item_name or 'WeChat' in item_name or 'wechat' in item_name.lower():
        return 'wechat'
    elif '现金' in item_name or 'cash' in item_name.lower():
        return 'cash'
    return 'cash'


def parse_excel_file(file_path, valid_categories=None, bank_accounts=None):
    """
    解析Excel文件，返回交易记录列表
    规则:
    1. 第二列支持"收"/"支"/"调"，其余全部忽略
    2. 不再检查黄色填充行（v8.1取消）
    3. 列结构:
       第1列: 日期
       第2列: 收/支/调
       第3列: 对象(收/支)或调出账户(调)
       第4列: 收费大类(收/支)或资金调拨(调)
       第5列: 具体项目(收/支)或转入账户(调)
       第6列: 应收金额(收/支)或计划调拨金额(调)
       第7列: 实收金额(收/支)或实际到账金额(调)
       第8列: 备注
    valid_categories: 从数据库读取的有效收费大类集合
    bank_accounts: 从数据库读取的银行账号名称集合，用于调拨匹配
    v8.3变更:
    - 支持"调"类型（资金调拨）
    - 调拨行校验银行账号匹配
    """
    records = []
    errors = []

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'success': False, 'error': f'无法读取Excel文件: {str(e)}'}

    VALID_CATEGORIES = valid_categories if valid_categories is not None else set()
    BANK_ACCOUNTS = bank_accounts if bank_accounts is not None else set()

    total_rows = 0
    skipped_other = 0
    imported = 0

    check_report = {
        'empty_counterparty': [],
        'empty_category': [],
        'invalid_category': [],
        'invalid_bank_from': [],   # v8.3 调出账户不匹配
        'invalid_bank_to': [],     # v8.3 转入账户不匹配
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        total_rows += 1

        cells = list(row)
        if len(cells) < 6:
            continue

        date_val = cells[0].value
        action_val = cells[1].value
        col3_val = cells[2].value if len(cells) > 2 else ''
        col4_val = cells[3].value if len(cells) > 3 else ''
        col5_val = cells[4].value if len(cells) > 4 else ''
        col6_val = cells[5].value if len(cells) > 5 else 0
        col7_val = cells[6].value if len(cells) > 6 else 0
        remark_val = cells[7].value if len(cells) > 7 else ''

        action_str = str(action_val).strip() if action_val else ''

        # v8.3: 支持收/支/调三种类型
        if action_str == '收':
            trans_type = 'income'
        elif action_str == '支':
            trans_type = 'expense'
        elif action_str == '调':
            trans_type = 'transfer'
        else:
            skipped_other += 1
            continue

        # 解析日期
        trans_date = parse_date(date_val)
        if not trans_date:
            errors.append(f'第{row_idx}行: 日期格式错误 [{date_val}]')
            continue

        # ===== 按类型分别处理 =====
        if trans_type == 'transfer':
            # 调拨行处理
            from_account = str(col3_val).strip() if col3_val else ''
            category_str = str(col4_val).strip() if col4_val else ''
            to_account = str(col5_val).strip() if col5_val else ''
            amount_planned = parse_amount(col6_val)
            amount_real = parse_amount(col7_val)
            remark = str(remark_val).strip() if remark_val else ''

            # 校验调出账户
            is_valid_from = True
            if from_account and BANK_ACCOUNTS and from_account not in BANK_ACCOUNTS:
                check_report['invalid_bank_from'].append({
                    'row': row_idx, 'value': from_account
                })
                is_valid_from = False

            # 校验转入账户
            is_valid_to = True
            if to_account and BANK_ACCOUNTS and to_account not in BANK_ACCOUNTS:
                check_report['invalid_bank_to'].append({
                    'row': row_idx, 'value': to_account
                })
                is_valid_to = False

            # 校验列4必须是"资金调拨"
            is_valid_cat = True
            if category_str and category_str != '资金调拨' and VALID_CATEGORIES:
                check_report['invalid_category'].append({
                    'row': row_idx, 'value': category_str
                })
                is_valid_cat = False

            records.append({
                'trans_date': trans_date,
                'trans_type': 'transfer',
                'from_account': from_account,
                'to_account': to_account,
                'category': category_str or '资金调拨',
                'amount_planned': amount_planned,
                'amount_real': amount_real,
                'amount': amount_planned,  # 兼容字段
                'remark': remark,
                '_valid_bank_from': is_valid_from,
                '_valid_bank_to': is_valid_to,
                '_valid_category': is_valid_cat,
            })
        else:
            # 收入/支出行处理
            counterparty_str = str(col3_val).strip() if col3_val else ''
            category_str = str(col4_val).strip() if col4_val else ''
            item_name = str(col5_val).strip() if col5_val else ''
            amount_receivable = parse_amount(col6_val)
            amount_real = parse_amount(col7_val)
            remark = str(remark_val).strip() if remark_val else ''

            if counterparty_str == '':
                check_report['empty_counterparty'].append(row_idx)
            if category_str == '':
                check_report['empty_category'].append(row_idx)

            is_valid_category = True
            if category_str != '' and VALID_CATEGORIES and category_str not in VALID_CATEGORIES:
                check_report['invalid_category'].append({
                    'row': row_idx, 'value': category_str
                })
                is_valid_category = False

            payment_method = detect_payment_method(item_name)

            records.append({
                'trans_date': trans_date,
                'trans_type': trans_type,
                'counterparty': counterparty_str,
                'category': category_str,
                'item_name': item_name,
                'amount_receivable': amount_receivable,
                'amount_real': amount_real,
                'payment_method': payment_method,
                'remark': remark,
                '_valid_category': is_valid_category,
            })
        imported += 1

    return {
        'success': True,
        'records': records,
        'stats': {
            'total_rows': total_rows,
            'skipped_yellow': 0,
            'skipped_other': skipped_other,
            'imported': imported,
            'errors': len(errors)
        },
        'check_report': check_report,
        'errors': errors
    }


def import_to_database(records, source_file=''):
    """将记录导入数据库 - v8.3 按类型写入对应表（收入/支出/调拨）"""
    from database import get_db_connection, generate_link_no

    conn = get_db_connection()
    cursor = conn.cursor()

    imported_count = 0
    try:
        for record in records:
            trans_type = record.get('trans_type', 'income')
            if trans_type == 'transfer':
                # v8.3: 调拨记录写入 transfer_records
                link_no = generate_link_no(cursor, 'TF')
                cursor.execute('''
                    INSERT INTO transfer_records
                    (trans_date, from_account, to_account, from_bank_id, to_bank_id,
                     amount_planned, amount_real, amount, remark, link_no)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    record['trans_date'],
                    record.get('from_account', ''),
                    record.get('to_account', ''),
                    record.get('from_bank_id'),
                    record.get('to_bank_id'),
                    record.get('amount_planned', 0),
                    record.get('amount_real', 0),
                    record.get('amount_planned', 0),
                    record.get('remark', ''),
                    link_no
                ))
            else:
                table = 'income_records' if trans_type == 'income' else 'expense_records'
                link_no = generate_link_no(cursor)  # v8.3.1: 收入/支出也生成link_no
                cursor.execute(f'''
                    INSERT INTO {table}
                    (trans_date, counterparty, category, item_name, amount_receivable, amount_real, payment_method, remark, source_file, link_no)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    record['trans_date'],
                    record.get('counterparty', ''),
                    record.get('category', ''),
                    record.get('item_name', ''),
                    record.get('amount_receivable', 0),
                    record.get('amount_real', 0),
                    record.get('payment_method', 'cash'),
                    record.get('remark', ''),
                    source_file,
                    link_no
                ))
            imported_count += 1

        conn.commit()
        return {'success': True, 'imported': imported_count}
    except Exception as e:
        conn.rollback()
        return {'success': False, 'error': str(e)}
    finally:
        conn.close()
