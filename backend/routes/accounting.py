"""
财务报表核心路由 - v8.5.0
功能：期初余额录入检测、科目管理、余额保存、Excel导入
"""
import os
import io
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from database import get_db_connection

accounting_bp = Blueprint('accounting', __name__)

# ========== 工具函数 ==========

def _current_year():
    return datetime.now().year


def _category_type(l1_code):
    """根据一级科目代码判断类别"""
    code = str(l1_code)
    if code.startswith('1') or code.startswith('2'):
        if code.startswith('1'):
            return 'asset'
        return 'liability'
    if code.startswith('4'):
        return 'equity'
    if code.startswith('5'):
        return 'cost'
    if code.startswith('6'):
        # 6开头：收入或费用
        code_int = int(code[:4]) if code[:4].isdigit() else 0
        if code_int >= 6001 and code_int <= 6111:
            return 'income'  # 收入类
        return 'expense'  # 费用类
    return 'other'


# ========== API: 检查是否需要录入期初余额 ==========

@accounting_bp.route('/check-opening', methods=['GET'])
def check_opening_balance():
    """
    检查是否需要录入期初余额：
    - 当年1月起的凭证不存在 → 检查是否已有期初余额
    - 24小时内跳过的不弹
    """
    year = _current_year()
    conn = get_db_connection()
    cursor = conn.cursor()

    # 1. 检查当年1月起的凭证是否存在
    cursor.execute('''
        SELECT COUNT(*) as cnt FROM vouchers
        WHERE YEAR(voucher_date) = %s AND MONTH(voucher_date) >= 1
          AND status = 'confirmed'
    ''', (year,))
    has_vouchers = cursor.fetchone()['cnt'] > 0

    # 2. 检查是否已有期初余额记录
    cursor.execute('''
        SELECT COUNT(*) as cnt FROM subject_balances sb
        JOIN accounting_periods ap ON sb.period_id = ap.id
        WHERE ap.year = %s AND ap.month = 1
    ''', (year,))
    has_balances = cursor.fetchone()['cnt'] > 0

    # 3. 检查24小时内是否跳过
    cursor.execute('''
        SELECT COUNT(*) as cnt FROM opening_balance_skip_log
        WHERE skipped_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)
    ''')
    recently_skipped = cursor.fetchone()['cnt'] > 0

    conn.close()

    required = not has_vouchers and not has_balances and not recently_skipped

    return jsonify({
        'code': 200,
        'required': required,
        'year': year,
        'has_vouchers': has_vouchers,
        'has_balances': has_balances,
        'recently_skipped': recently_skipped,
    })


# ========== API: 获取科目列表（用于录入） ==========

@accounting_bp.route('/subjects', methods=['GET'])
def get_subjects_for_opening():
    """
    返回所有一级科目 + 已有的二级科目。
    用于期初余额录入界面的下拉选择和表格显示。
    """
    conn = get_db_connection()
    cursor = conn.cursor()

    # 一级科目：按编码排序
    cursor.execute('''
        SELECT code, name, category, direction
        FROM account_subjects_l1
        ORDER BY code
    ''')
    l1_list = cursor.fetchall()

    # 二级科目：带 parent_code
    cursor.execute('''
        SELECT l2.code, l2.name, l1.code as parent_code, l1.name as parent_name
        FROM account_subjects_l2 l2
        JOIN account_subjects_l1 l1 ON l2.parent_id = l1.id
        WHERE l2.is_active = 1
        ORDER BY l1.code, l2.code
    ''')
    l2_list = cursor.fetchall()

    conn.close()

    # 按一级科目分组
    grouped = []
    for l1 in l1_list:
        l2_children = [{
            'code': r['code'],
            'name': r['name'],
        } for r in l2_list if r['parent_code'] == l1['code']]

        grouped.append({
            'code': l1['code'],
            'name': l1['name'],
            'category': l1['category'],
            'direction': l1['direction'],
            'children': l2_children,
        })

    return jsonify({
        'code': 200,
        'data': {
            'subjects': grouped,
        }
    })


# ========== API: 保存期初余额 ==========

@accounting_bp.route('/opening-balance', methods=['POST'])
def save_opening_balance():
    """
    保存期初余额：
    1. 创建会计期间（如不存在）
    2. 保存余额记录
    3. 资产负债平衡校验
    4. 自动生成年初余额凭证
    """
    data = request.get_json() or {}
    year = data.get('year', _current_year())
    entries = data.get('entries', [])  # [{l1_code, l2_code|null, balance}]

    if not entries:
        return jsonify({'code': 400, 'message': '请至少录入一个科目的余额'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # 1. 创建或获取1月会计期间
        cursor.execute('SELECT id FROM accounting_periods WHERE year = %s AND month = 1', (year,))
        row = cursor.fetchone()
        if row:
            period_id = row['id']
        else:
            # 批量创建全年12个月
            for m in range(1, 13):
                cursor.execute(
                    'INSERT IGNORE INTO accounting_periods (year, month, is_year_end) VALUES (%s, %s, %s)',
                    (year, m, 1 if m == 12 else 0)
                )
            cursor.execute('SELECT id FROM accounting_periods WHERE year = %s AND month = 1', (year,))
            period_id = cursor.fetchone()['id']

        # 2. 清空已有余额（支持重新录入）
        cursor.execute('DELETE FROM subject_balances WHERE period_id = %s', (period_id,))

        # 3. 插入新余额
        # 先获取所有一级科目的 direction
        cursor.execute('SELECT code, direction, name, category FROM account_subjects_l1')
        l1_info = {r['code']: r for r in cursor.fetchall()}

        # 分类汇总（用于校验和凭证生成）
        assets_total = 0
        liabilities_total = 0
        equity_total = 0

        for entry in entries:
            l1_code = str(entry.get('l1_code', '')).strip()
            l2_code = entry.get('l2_code') or None
            balance = float(entry.get('balance', 0) or 0)
            is_l1 = entry.get('is_l1_entry', False) or (l2_code is None)

            if not l1_code:
                continue

            info = l1_info.get(l1_code, {})
            direction = info.get('direction', 'debit')
            category = info.get('category', '')

            # 根据方向确定借贷方向的影响
            # debit方向科目：正余额 = 借方余额
            # credit方向科目：正余额 = 贷方余额
            opening = balance

            cursor.execute('''
                INSERT INTO subject_balances
                (period_id, subject_l1_code, subject_l2_code, opening_balance, is_l1_entry)
                VALUES (%s, %s, %s, %s, %s)
            ''', (period_id, l1_code, l2_code, opening, is_l1))

            # 汇总用于校验
            cat_type = _category_type(l1_code)
            if cat_type == 'asset':
                assets_total += opening
            elif cat_type == 'liability':
                liabilities_total += opening
            elif cat_type == 'equity':
                equity_total += opening
            # 损益类不参与资产负债平衡校验（年初余额为0）

        # 4. 资产负债平衡校验（强制阻断）
        total_liab_eq = liabilities_total + equity_total
        diff = round(assets_total - total_liab_eq, 2)

        if abs(diff) > 0.01:
            conn.rollback()
            return jsonify({
                'code': 400,
                'message': f'资产负债不平衡！资产总计 {assets_total:.2f} ≠ 负债+权益 {total_liab_eq:.2f}，差额 {diff:.2f}',
                'data': {
                    'assets': assets_total,
                    'liabilities': liabilities_total,
                    'equity': equity_total,
                    'difference': diff,
                }
            }), 400

        # 5. 自动生成年初余额凭证（Q6: 是）
        _generate_opening_voucher(cursor, period_id, year, entries, l1_info)

        conn.commit()
        return jsonify({
            'code': 200,
            'message': '期初余额录入成功，已生成年初余额凭证',
            'data': {
                'assets': assets_total,
                'liabilities': liabilities_total,
                'equity': equity_total,
            }
        })

    except Exception as e:
        conn.rollback()
        return jsonify({'code': 500, 'message': f'保存失败: {str(e)}'}), 500
    finally:
        conn.close()


def _generate_opening_voucher(cursor, period_id, year, entries, l1_info):
    """根据期初余额自动生成年初余额凭证"""
    from database import generate_voucher_no

    # 生成凭证
    voucher_no = f'QC-{year}'
    # 检查是否已有同号凭证，有则删除重建
    cursor.execute('SELECT id FROM vouchers WHERE voucher_no = %s', (voucher_no,))
    existing = cursor.fetchone()
    if existing:
        cursor.execute('DELETE FROM voucher_entries WHERE voucher_id = %s', (existing['id'],))
        cursor.execute('DELETE FROM vouchers WHERE id = %s', (existing['id'],))

    voucher_date = f'{year}-01-01'
    cursor.execute('''
        INSERT INTO vouchers (voucher_no, voucher_date, source_type, source_table, status, remark)
        VALUES (%s, %s, 'opening', 'subject_balances', 'confirmed', '年初余额录入自动生成')
    ''', (voucher_no, voucher_date))
    voucher_id = cursor.lastrowid

    seq = 1
    for entry in entries:
        l1_code = str(entry.get('l1_code', '')).strip()
        l2_code = entry.get('l2_code') or None
        balance = float(entry.get('balance', 0) or 0)

        if not l1_code or balance == 0:
            continue

        info = l1_info.get(l1_code, {})
        direction = info.get('direction', 'debit')
        l1_name = info.get('name', '')

        # 确定借贷方向
        # debit方向科目（资产、费用）：余额在借方 → 借记该科目
        # credit方向科目（负债、权益、收入）：余额在贷方 → 贷记该科目
        if direction == 'debit':
            entry_direction = 'debit'
        else:
            entry_direction = 'credit'

        subject_name = l1_name
        if l2_code:
            # 查询二级科目名称
            cursor.execute('SELECT name FROM account_subjects_l2 WHERE code = %s', (l2_code,))
            l2_row = cursor.fetchone()
            if l2_row:
                subject_name = l2_row['name']

        cursor.execute('''
            INSERT INTO voucher_entries
            (voucher_id, seq_no, subject_l1_code, subject_l2_code, subject_name, direction, amount, summary)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (voucher_id, seq, l1_code, l2_code, subject_name, entry_direction, abs(balance), '年初余额'))
        seq += 1


# ========== API: 跳过期初余额录入 ==========

@accounting_bp.route('/opening-balance/skip', methods=['POST'])
def skip_opening_balance():
    """记录跳过时间，24小时内不再弹出"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO opening_balance_skip_log (skipped_at) VALUES (NOW())')
    conn.commit()
    conn.close()
    return jsonify({'code': 200, 'message': '已跳过，24小时内不再提醒'})


# ========== API: 下载Excel模板 ==========

@accounting_bp.route('/opening-balance/template', methods=['GET'])
def download_template():
    """下载期初余额录入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = '期初余额录入'

    # 标题行
    headers = ['一级科目编码', '一级科目名称', '二级科目编码（可选）', '二级科目名称（可选）', '年初余额']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 示例数据
    examples = [
        ['1001', '库存现金', '', '', '5000.00'],
        ['1002', '银行存款', '1002.01', '银行存款—对公账户', '150000.00'],
        ['1002', '银行存款', '1002.02', '银行存款—招商银行', '80000.00'],
        ['1601', '固定资产', '', '', '500000.00'],
        ['2202', '应付账款', '', '', '0'],
        ['4104', '利润分配', '', '', '200000.00'],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # 说明 sheet
    ws_help = wb.create_sheet('填写说明')
    help_text = [
        '填写说明：',
        '1. 一级科目编码和名称必须填写，请从法定科目表中选择',
        '2. 二级科目为可选：如果不填则表示直接在一级科目下录入余额',
        '3. 如果填写了二级科目，则一级科目下的余额由二级科目汇总（不能同时录入一级和二级）',
        '4. 年初余额列填写数字，不需要逗号分隔符',
        '5. 损益类科目（收入、费用）年初余额为0，可不填',
        '6. 导入前请确保：资产总计 = 负债总计 + 所有者权益总计',
    ]
    for row_idx, text in enumerate(help_text, 1):
        ws_help.cell(row=row_idx, column=1, value=text)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='期初余额录入模板.xlsx'
    )


# ========== API: 导入Excel ==========

@accounting_bp.route('/opening-balance/import', methods=['POST'])
def import_opening_balance():
    """从Excel导入期初余额"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件'}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'code': 400, 'message': '请上传 .xlsx 文件'}), 400

    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb[wb.sheetnames[0]]

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            l1_code = str(row[0]).strip()
            l1_name = str(row[1]).strip() if row[1] else ''
            l2_code = str(row[2]).strip() if row[2] else None
            l2_name = str(row[3]).strip() if row[3] else ''
            balance_str = str(row[4]).strip() if row[4] else '0'

            try:
                balance = float(balance_str.replace(',', ''))
            except ValueError:
                balance = 0

            entries.append({
                'l1_code': l1_code,
                'l1_name': l1_name,
                'l2_code': l2_code or None,
                'l2_name': l2_name,
                'balance': balance,
                'is_l1_entry': not l2_code,
            })

        # 校验
        if not entries:
            return jsonify({'code': 400, 'message': '未读取到有效数据'}), 400

        return jsonify({
            'code': 200,
            'message': f'成功读取 {len(entries)} 条记录，请确认后保存',
            'data': {
                'entries': entries,
                'count': len(entries),
            }
        })

    except Exception as e:
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


# ========== API: 获取已录入的期初余额 ==========

@accounting_bp.route('/opening-balance', methods=['GET'])
def get_opening_balance():
    """获取已录入的期初余额（用于编辑/查看）"""
    year = request.args.get('year', _current_year(), type=int)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT sb.*, ap.year, ap.month
        FROM subject_balances sb
        JOIN accounting_periods ap ON sb.period_id = ap.id
        WHERE ap.year = %s AND ap.month = 1
        ORDER BY sb.subject_l1_code, sb.subject_l2_code
    ''', (year,))
    rows = cursor.fetchall()
    conn.close()

    entries = [{
        'l1_code': r['subject_l1_code'],
        'l2_code': r['subject_l2_code'],
        'balance': float(r['opening_balance']),
        'is_l1_entry': bool(r['is_l1_entry']),
    } for r in rows]

    return jsonify({
        'code': 200,
        'data': {
            'year': year,
            'entries': entries,
        }
    })
